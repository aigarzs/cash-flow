from datetime import date, datetime

from PyQt6.QtCore import Qt, QModelIndex
from PyQt6.QtGui import QBrush, QColor
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QDateEdit, QLabel, QComboBox, QPushButton, QFileDialog, \
    QTabWidget, QCheckBox, QApplication, QStackedWidget

import pandas as pd
import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
import mplcursors


from cash_flow.ui.AWidgets import ATable, ATableModel
from cash_flow.ui.Browse_ActualPayment import Browse_ActualPayment
from cash_flow.ui.Browse_ActualReceipt import Browse_ActualReceipt
from cash_flow.ui.Browse_BudgetedPayment import Browse_BudgetedPayment
from cash_flow.ui.Browse_BudgetedReceipt import Browse_BudgetedReceipt
from cash_flow.ui.Browse_PendingPayment import Browse_PendingPayment
from cash_flow.ui.Browse_PendingReceipt import Browse_PendingReceipt
from cash_flow.util.Converters import date_format

from Browse_ActualReceipt import Browse_ActualReceipt


class CashFlowReport(QWidget):

    def __init__(self, engine):
        super().__init__(None)
        self.engine = engine

        vbox = QVBoxLayout()

        filterbox_from = QHBoxLayout()
        filterbox_through = QHBoxLayout()
        filterbox_freq = QHBoxLayout()
        label_filter = QLabel("Filtrs")
        label_filter.setStyleSheet("font-weight: bold")
        label_dateFrom = QLabel("No datuma")
        self.filter_dateFrom = QDateEdit()
        self.filter_dateFrom.setCalendarPopup(True)
        self.filter_dateFrom.setDisplayFormat("dd.MMMM.yyyy")
        self.filter_dateFrom.setDate(date(datetime.now().year, 1, 1))
        self.filter_dateFrom.dateChanged.connect(self.requery)
        filterbox_from.addWidget(label_dateFrom)
        filterbox_from.addWidget(self.filter_dateFrom)
        label_dateThrough = QLabel("Līdz datumam")
        self.filter_dateThrough = QDateEdit()
        self.filter_dateThrough.setCalendarPopup(True)
        self.filter_dateThrough.setDisplayFormat("dd.MMMM.yyyy")
        self.filter_dateThrough.setDate(datetime.now())
        self.filter_dateThrough.dateChanged.connect(self.requery)
        filterbox_through.addWidget(label_dateThrough)
        filterbox_through.addWidget(self.filter_dateThrough)
        label_freq = QLabel("Periodiskums")
        self.freq_map = {
            "Diena": "day",
            "Nedēļa": "week",
            "Mēnesis": "month",
            "Kvartāls": "quarter",
            "Gads": "year"
        }
        self.filter_Frequency = QComboBox(self)
        self.filter_Frequency.addItems(self.freq_map.keys())
        self.filter_Frequency.setCurrentText("Mēnesis")
        self.filter_Frequency.currentIndexChanged.connect(self.requery)
        filterbox_freq.addWidget(label_freq)
        filterbox_freq.addWidget(self.filter_Frequency)

        table_widget = QWidget()
        table_box = QVBoxLayout()
        table_toolbox = QHBoxLayout()
        btn_excel_cashflow = QPushButton("Eksportēt uz Excel")
        btn_excel_cashflow.clicked.connect(self.export_to_excel_cashflow)
        table_toolbox.addStretch()
        table_toolbox.addWidget(btn_excel_cashflow)
        self.table = ATable(self)
        self.table.setModel(CashFlowReportModel(self.table, self.engine))
        table_box.addLayout(table_toolbox)
        table_box.addWidget(self.table)
        table_widget.setLayout(table_box)

        graph_widget = QWidget()
        graph_layout = QVBoxLayout(self)

        # Add export button
        graph_toolbox = QHBoxLayout()
        label_legend = QLabel("Rādīt leģendu")
        check_legend = QCheckBox()
        check_legend.setChecked(True)
        self.show_legend = True
        check_legend.stateChanged.connect(self.toggle_legend)
        btn_jpg = QPushButton("Eksportēt uz JPG")
        btn_jpg.clicked.connect(self.export_to_jpg)
        graph_toolbox.addWidget(label_legend)
        graph_toolbox.addWidget(check_legend)
        graph_toolbox.addStretch()
        graph_toolbox.addWidget(btn_jpg)
        self.canvas = FigureCanvas(plt.Figure(figsize=(12, 6)))
        self.cursor_bars = None
        # self.cursor_lines = None
        self.bar_artists = []
        self.annotations = []
        self.canvas.mpl_connect("button_press_event", self.hide_tooltips)

        graph_layout.addLayout(graph_toolbox)
        graph_layout.addWidget(self.canvas)
        graph_widget.setLayout(graph_layout)

        self.checkreport_stacked = QStackedWidget()

        self.checkreport_widget = QWidget()
        checkreport_box = QVBoxLayout()
        checkreport_toolbox = QHBoxLayout()
        btn_browse = QPushButton("Izvērst")
        btn_browse.clicked.connect(self.clicked_browse)
        btn_excel_checkreport = QPushButton("Eksportēt uz Excel")
        btn_excel_checkreport.clicked.connect(self.export_to_excel_checkreport)
        checkreport_toolbox.addStretch()
        checkreport_toolbox.addWidget(btn_browse)
        checkreport_toolbox.addWidget(btn_excel_checkreport)
        self.checkreport = CheckingReportTable(self.requery, self)
        self.checkreport.setModel(CheckingReportModel(self.checkreport, self.engine, self.table.model()))
        self.checkreport.doubleClicked.connect(self.browse_checkreport)
        checkreport_box.addLayout(checkreport_toolbox)
        checkreport_box.addWidget(self.checkreport)
        self.checkreport_widget.setLayout(checkreport_box)

        self.browse_actual_receipt = Browse_ActualReceipt(self.engine, self.return_checkreport)
        self.browse_actual_payment = Browse_ActualPayment(self.engine, self.return_checkreport)
        self.browse_pending_receipt = Browse_PendingReceipt(self.engine, self.return_checkreport)
        self.browse_pending_payment = Browse_PendingPayment(self.engine, self.return_checkreport)
        self.browse_budgeted_receipt = Browse_BudgetedReceipt(self.engine, self.return_checkreport)
        self.browse_budgeted_payment = Browse_BudgetedPayment(self.engine, self.return_checkreport)

        self.checkreport_stacked.addWidget(self.checkreport_widget)
        self.checkreport_stacked.addWidget(self.browse_actual_receipt)
        self.checkreport_stacked.addWidget(self.browse_actual_payment)
        self.checkreport_stacked.addWidget(self.browse_pending_receipt)
        self.checkreport_stacked.addWidget(self.browse_pending_payment)
        self.checkreport_stacked.addWidget(self.browse_budgeted_receipt)
        self.checkreport_stacked.addWidget(self.browse_budgeted_payment)

        tabs = QTabWidget()
        tabs.addTab(table_widget, "Naudas plūsmas atskaite")
        tabs.addTab(graph_widget, "Naudas plūsmas grafiks")
        tabs.addTab(self.checkreport_stacked, "Pārbaudes atskaite")

        vbox.addWidget(label_filter)
        vbox.addLayout(filterbox_from)
        vbox.addLayout(filterbox_through)
        vbox.addLayout(filterbox_freq)
        vbox.addWidget(tabs)
        self.setLayout(vbox)
        self.requery()

    def requery(self):
        selected_freq = self.filter_Frequency.currentText()
        freq = self.freq_map[selected_freq]
        self.table.model().set_filter({"date from": self.filter_dateFrom.date().toPyDate(),
                                       "date through": self.filter_dateThrough.date().toPyDate(),
                                       "frequency": freq})
        self.plot_cashflow()
        self.checkreport.model().requery()

    def return_checkreport(self):
        self.checkreport_stacked.setCurrentWidget(self.checkreport_widget)

    def clicked_browse(self):
        selected = self.checkreport.selectedIndexes()
        if selected:
            index = selected[0]
            self.browse_checkreport(index)

    def browse_checkreport(self, index):
        # print(index.row(), index.column())
        forms_map = [None,
                     None,
                     None,
                     self.browse_actual_receipt,
                     self.browse_actual_payment,
                     self.browse_pending_receipt,
                     self.browse_pending_payment,
                     self.browse_budgeted_receipt,
                     self.browse_budgeted_payment,
                     None,
                     None,
                     None,
                     None,
                     None]

        form = forms_map[index.column()]
        if form:
            form.table.model().set_filter({"Date_From": self.filter_dateFrom.date().toPyDate(),
                                           "Date_Through": self.filter_dateThrough.date().toPyDate(),
                                           "Period_End": self.checkreport.model().DATA.iloc[index.row(), 0],
                                           "Definition_ID": self.checkreport.model().DATA.iloc[index.row(), 1],
                                           "Frequency": self.freq_map.get(self.filter_Frequency.currentText(), "month")})
            self.checkreport_stacked.setCurrentWidget(form)


    def export_to_excel_checkreport(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if file_name:
            if not file_name.endswith(".xlsx"):
                file_name += ".xlsx"
            # Save DataFrame to Excel
            self.checkreport.model().DATA.to_excel(file_name, index=True)

    def export_to_excel_cashflow(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if file_name:
            if not file_name.endswith(".xlsx"):
                file_name += ".xlsx"
            # Save DataFrame to Excel
            self.table.model().DATA.to_excel(file_name, index=True)

            # Load workbook for formatting
            wb = load_workbook(file_name)
            ws = wb.active

            # Align entire first column (Index) to left
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")

            # Formatting row colors
            dark_blue_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
            dark_cyan_fill = PatternFill(start_color="008B8B", end_color="008B8B", fill_type="solid")
            # Define font for white text
            white_font = Font(color="FFFFFF")

            row = 1 # First row for column headers
            for def_type in self.table.model().FORMAT:
                row = row + 1
                if def_type == 2:  # Totals
                    for cell in ws[row]:
                        cell.fill = dark_blue_fill
                        cell.font = white_font
                elif def_type == 3:  # Balances
                    for cell in ws[row]:
                        cell.fill = dark_cyan_fill
                        cell.font = white_font

            wb.save(file_name)
            # print(f"Data exported to {file_name} with formatting.")

    def export_to_jpg(self):
        self.hide_tooltips()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Image", "", "JPG files (*.jpg);;All Files (*)")
        if file_path:
            original_facecolor = self.canvas.figure.get_facecolor()
            self.canvas.figure.set_facecolor('white')
            self.canvas.figure.savefig(file_path, format='jpg')
            self.canvas.figure.set_facecolor(original_facecolor)
            self.plot_cashflow()

    def toggle_legend(self, state):
        if state == Qt.CheckState.Checked.value:
            self.show_legend = True
        else:
            self.show_legend = False

        self.plot_cashflow()

    def hide_tooltips(self, event=None):
        for ann in self.annotations:
            ann.set_visible(False)
        self.annotations.clear()
        self.canvas.draw_idle()


    def plot_cashflow(self):
        # print("plot_cashflow()")
        cashflow = self.table.model().graph_pivot
        balances = self.table.model().graph_balances
        self.canvas.figure.clf()
        ax = self.canvas.figure.add_subplot(111)
        periods = cashflow.columns
        x = np.arange(len(periods))

        pos_bottom = np.zeros(len(periods))
        neg_bottom = np.zeros(len(periods))
        colors = plt.cm.tab20.colors

        self.bar_artists = []
        self.annotations = []

        for i, account in enumerate(cashflow.index):
            values = cashflow.loc[account].values
            bar_vals = []
            bar_bottom = []

            for j, val in enumerate(values):
                if val >= 0:
                    bar_bottom.append(pos_bottom[j])
                    pos_bottom[j] += val
                else:
                    bar_bottom.append(neg_bottom[j])
                    neg_bottom[j] += val
                bar_vals.append(val)

            bars = ax.bar(
                x,
                bar_vals,
                bottom=bar_bottom,
                label=account,
                color=colors[i % len(colors)]
            )

            for idx, bar in enumerate(bars):
                if bar_vals[idx] != 0:
                    self.bar_artists.append((bar, account, bar_vals[idx], periods[idx]))

        net_cashflow = cashflow.sum(axis=0).values
        line1, = ax.plot(x, net_cashflow, color='black', label='Neto naudas plūsma', marker='o', linewidth=2)

        line2, = ax.plot(x, balances, color='blue', label='Bilance', linestyle='--', marker='x',
                         linewidth=2)

        ax.set_xticks(x)
        ax.set_xticklabels([p.strftime(date_format()) for p in periods], rotation=25, ha='right')
        ax.set_ylabel("Summa BV")
        ax.set_title("Naudas plūsma pa periodiem")
        ax.grid(True)
        if self.show_legend:
            ax.legend()

        if self.cursor_bars:
            self.cursor_bars.remove()

        """
        if self.cursor_lines:
            self.cursor_lines.remove()
        """

        self.cursor_bars = mplcursors.cursor([b[0] for b in self.bar_artists], hover=True)

        @self.cursor_bars.connect("add")
        def on_add_bar(sel):
            bar, account, val, period = next((b for b in self.bar_artists if b[0] == sel.artist),
                                             (None, None, None, None))
            if bar:
                sel.annotation.set_text(f"{account}\n{period.strftime(date_format())}\n{val:,.2f}")
                sel.annotation.get_bbox_patch().set_facecolor(bar.get_facecolor())
                sel.annotation.get_bbox_patch().set_alpha(0.7)
                self.annotations.append(sel.annotation)

        """
        
        self.cursor_lines = mplcursors.cursor([line1, line2], hover=True)

        @self.cursor_lines.connect("add")
        def on_add_line(sel):
            label = line1.get_label() if sel.artist == line1 else line2.get_label()
            period = periods[int(sel.index)]
            y = sel.target[1]
            sel.annotation.set_text(f"{label}\n{period.strftime('%Y-%m-%d')}\n{y:,.2f}")
            sel.annotation.get_bbox_patch().set_facecolor((1, 1, 1, 0.8))
            self.annotations.append(sel.annotation)
        """

        self.canvas.draw()




class CashFlowReportModel(ATableModel):

    def _do_requery(self):
        engine = self.engine

        # Setup filters
        date_from = self.FILTER["date from"].strftime("%Y-%m-%d")
        date_through = self.FILTER["date through"].strftime("%Y-%m-%d")
        period = self.FILTER["frequency"]

        # Mapping period to offset (end of period)
        period_offsets = {
            "day": pd.offsets.Day(0),
            "week": pd.offsets.Week(weekday=6),  # Sunday
            "month": pd.offsets.MonthEnd(0),
            "quarter": pd.offsets.QuarterEnd(startingMonth=12),
            "year": pd.offsets.YearEnd(0)
        }
        date_offset = period_offsets.get(period, pd.offsets.MonthEnd(0))
        date_freq = {
            "day": "D",
            "week": "W-SUN",  # Sunday
            "month": "ME",
            "quarter": "QE",
            "year": "YE"
        }
        date_frequency = date_freq.get(period, "ME")

        # Load all data from database

        # handling sqlite issues with datetime timepart
        next_day_date_through = (pd.to_datetime(date_through) + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
        previous_day_date_from = (pd.to_datetime(date_from) - pd.Timedelta(days=1)).strftime('%Y-%m-%d')

        actual = pd.read_sql_query(
            'SELECT * FROM G09_CashFlow_Actual_Corresponding WHERE date < "' + next_day_date_through + '" ', engine)
        pending = pd.read_sql_query(
            'SELECT * FROM H02_CashFlow_Pending_Corresponding WHERE p_date < "' + next_day_date_through + '" ', engine)
        budgeted = pd.read_sql_query(
            'SELECT * FROM F01_BudgetEntries WHERE date > "' + previous_day_date_from + '" AND date < "' + next_day_date_through + '" ',
            engine)
        cash = pd.read_sql_query('SELECT * FROM G01_CashTransactions WHERE date < "' + next_day_date_through + '" ',
                                 engine)
        definition_df = pd.read_sql_table("E01_CashFlowDefinition", engine)

        definition_accounts_df = pd.read_sql_table("E01_CashFlowDefinitionAccounts", engine)
        definition_totals_df = pd.read_sql_table("E01_CashFlowDefinitionTotals", engine)

        # Ensure required columns and formats

        actual['period_end'] = pd.to_datetime(actual['date'], format='mixed', errors='coerce').apply(
            lambda d: date_offset.rollforward(d) if pd.notnull(d) else pd.NaT)
        actual['cf_amount'] = np.where(actual['entry_type'] == 'CR', actual['amount_LC'], -actual['amount_LC'])
        pending['period_end'] = pd.to_datetime(pending['p_date'], format='mixed', errors='coerce').apply(
            lambda d: date_offset.rollforward(d) if pd.notnull(d) else pd.NaT)
        pending['cf_amount'] = np.where(pending['entry_type'] == 'CR', pending['p_amount_LC'], -pending['p_amount_LC'])
        budgeted['period_end'] = pd.to_datetime(budgeted['date'], format='mixed', errors='coerce').apply(
            lambda d: date_offset.rollforward(d) if pd.notnull(d) else pd.NaT)
        budgeted['cf_amount'] = np.where(budgeted['cash_type'] == 'Receipt', budgeted['amount_LC'],
                                         -budgeted['amount_LC'])
        cash['period_end'] = pd.to_datetime(cash['date'], format='mixed', errors='coerce').apply(
            lambda d: date_offset.rollforward(d) if pd.notnull(d) else pd.NaT)
        cash['cf_amount'] = np.where(cash['entry_type'] == 'DR', cash['amount_LC'], -cash['amount_LC'])

        definition_df.rename(columns={"id": "definition_id"}, inplace=True)
        definition_acc_df = definition_df[definition_df["definition_type"] == 1]
        definition_tot_df = definition_df[definition_df["definition_type"] == 2]
        definition_bal_df = definition_df[definition_df["definition_type"] == 3]

        # -------------------------------------------------------------------
        # ************** (1) CashFlow based on accounts definitions *********
        # -------------------------------------------------------------------

        #  Prepare chart of CF definitions

        # Merge accounts into accounts definition
        definition_accounts_df = pd.merge(
            definition_acc_df,
            definition_accounts_df,
            on="definition_id", how="left")
        definition_accounts_df.drop(columns=["id"], inplace=True)

        actual = pd.merge(definition_accounts_df,
                          actual,
                          left_on=['cash_type', 'account'],  # Columns in definition_df
                          right_on=['cash_type', 'account'],  # Corresponding columns in transactions_df
                          how='left'
                          )

        pending = pd.merge(definition_accounts_df,
                           pending,
                           left_on=['cash_type', 'account'],  # Columns in definition_df
                           right_on=['cash_type', 'account'],  # Corresponding columns in transactions_df
                           how='left'
                           )

        budgeted = pd.merge(definition_acc_df,
                            budgeted,
                            left_on=['definition_id'],  # Columns in definition_df
                            right_on=['definition_id'],  # Corresponding columns in transactions_df
                            how='left'
                            )

        # -------------------
        # GROUP & MERGE
        # -------------------
        def get_period_totals(df, label):
            grouped = df.groupby(['definition_id', 'period_end', 'cash_type'])['cf_amount'].sum().unstack(fill_value=0)
            columns = ["Receipt", "Payment"]
            grouped = grouped.reindex(columns=columns, fill_value=0)
            grouped.columns = [f"{label}_{col}" for col in grouped.columns]
            return grouped

        actual_period = get_period_totals(actual, 'Actual')
        pending_period = get_period_totals(pending, 'Pending')
        budgeted_period = get_period_totals(budgeted, 'Budgeted')

        # Combine all
        combined = actual_period.join(pending_period, how='outer') \
            .join(budgeted_period, how='outer') \
            .fillna(0).reset_index()

        # -------------------
        # SPLIT PAST / FUTURE
        # -------------------
        today = pd.to_datetime(date.today())
        this_period_end = date_offset.rollforward(today)

        # Split
        past = combined[combined['period_end'] < this_period_end]
        future = combined[combined['period_end'] >= this_period_end]

        # -------------------
        # CASHFLOW LOGIC
        # -------------------

        # For past: use only actuals
        past['income'] = past.get('Actual_Receipt', 0)
        past['expense'] = past.get('Actual_Payment', 0)

        # For future: use max(budgeted, actual+pending)
        future['actual_plus_pending_income'] = future.get('Actual_Receipt', 0) + future.get('Pending_Receipt', 0)
        future['actual_plus_pending_expense'] = future.get('Actual_Payment', 0) + future.get('Pending_Payment', 0)

        future['income'] = future[['Budgeted_Receipt', 'actual_plus_pending_income']].max(axis=1)
        future['expense'] = future[['Budgeted_Payment', 'actual_plus_pending_expense']].min(axis=1)

        # Combine
        cashflow = pd.concat([past, future], ignore_index=True)
        cashflow['net_cashflow'] = cashflow['income'] + cashflow['expense']

        # Create and store cashflow checking report

        self.checking_report = pd.merge(definition_acc_df, cashflow, on="definition_id", how="right")
        self.checking_report.sort_values(["period_end", "key"], inplace=True)
        self.checking_report.drop(columns=["key", "definition_type"], inplace=True)
        self.checking_report = self.checking_report.reindex(
            columns=["period_end", "definition_id", "name", "Actual_Receipt", "Actual_Payment", "Pending_Receipt", "Pending_Payment",
                     "Budgeted_Receipt", "Budgeted_Payment", "actual_plus_pending_income",
                     "actual_plus_pending_expense",
                     "income", "expense", "net_cashflow"])
        self.checking_report = self.checking_report[self.checking_report["period_end"] >= pd.to_datetime(date_from)]

        # -------------------
        # CASHFLOW OUTPUT
        # -------------------

        # Pivot first
        pivot_all_periods = cashflow.pivot_table(
            index='definition_id',
            columns='period_end',
            values='net_cashflow',
            aggfunc='sum'
        ).fillna(0)

        # Fill missing periods and drop unnecessary

       # Build full period range
        min_period = date_offset.rollforward(pd.to_datetime(date_from))
        max_period = date_offset.rollforward(pd.to_datetime(date_through))

        report_periods = pd.date_range(start=min_period, end=max_period, freq=date_frequency)

        # Reindex pivot to include all requested periods
        pivot_cf = pivot_all_periods.reindex(columns=report_periods, fill_value=0)

        # Sort columns just in case
        pivot_cf = pivot_cf.sort_index(axis=1)

        # Store pivot_cf for cash flow figure
        self.graph_pivot = pd.merge(definition_acc_df, pivot_cf, on="definition_id", how="right")
        self.graph_pivot.drop(columns=["definition_id", "key", "definition_type"], inplace=True)
        self.graph_pivot = self.graph_pivot.set_index("name")

        # Ensure 0 instead of NaN in empty rows
        pivot_cf = pd.merge(definition_acc_df["definition_id"], pivot_cf, left_on="definition_id",
                            right_on="definition_id", how="left").fillna(0)
        pivot_cf.set_index("definition_id", inplace=True)

        # -----------------------------------------------------------------------------------------------------------------
        # ********************** (2) Start working with totals based on totaled definitions *******************************
        # -----------------------------------------------------------------------------------------------------------------

        # Merge totals into totals definition
        definition_totals_df = pd.merge(
            definition_tot_df,
            definition_totals_df,
            on="definition_id", how="left")
        definition_totals_df.drop(columns=["id"], inplace=True)

        # Merge totals definition with summarized accounts pivot

        merged_totals = pd.merge(definition_totals_df,
                                 pivot_cf,
                                 left_on='definition_summarized',
                                 right_on='definition_id',
                                 how='left'
                                 )

        # Drop unnecessary columns

        merged_totals.drop(columns=['key', 'definition_type', 'name', 'definition_summarized'], inplace=True)

        # Choose value_columns for further summarization

        value_columns = [col for col in merged_totals.columns if col not in ['definition_id']]

        # Summarize value_columns based on group value

        summarized_totals = merged_totals.groupby('definition_id', group_keys=False)[value_columns].sum()

        # .reset_index()

        summarized_totals.fillna(0, inplace=True)

        # ------------------------------------------------------------------------------------------------------------
        # ********************** (3) Start working with balances on end of each period *******************************
        # ------------------------------------------------------------------------------------------------------------

        # Step 1. Determine periods
        today = pd.to_datetime(date.today())
        this_period_end = date_offset.rollforward(today)
        min_period = date_offset.rollforward(pd.to_datetime(date_from))
        max_period = date_offset.rollforward(pd.to_datetime(date_through))

        report_periods = pd.date_range(start=min_period, end=max_period, freq=date_frequency)

        # Step 2. Determine past balances

        # Group and sort actual bank balance by period
        actual_cash_by_period = cash.groupby('period_end')['cf_amount'].sum().sort_index()
        actual_cumulative_cash = actual_cash_by_period.cumsum()

        if max_period >= this_period_end:
            past_balances = actual_cumulative_cash[actual_cumulative_cash.index < this_period_end]
        else:
            past_balances = actual_cumulative_cash[actual_cumulative_cash.index <= max_period]

        # Step 3. Determine future balances

        if min_period >= this_period_end:
            future_periods = pd.date_range(start=this_period_end, end=max_period, freq=date_frequency)
        else:
            future_periods = report_periods[report_periods >= this_period_end]

        last_actual_balance = past_balances.iloc[-1] if not past_balances.empty else 0

        # Future cashflows from pivot_all_periods
        cashflow_by_period = pivot_all_periods.sum(axis=0)

        bal = last_actual_balance
        future_balances = {}
        for period in future_periods:
            bal += cashflow_by_period.get(period, 0)
            future_balances[period] = bal

        # Step 4. Combine past and future into all closing balance series, adjust periods
        all_closing_balances = pd.Series(dtype=float)
        all_closing_balances = pd.concat([
            past_balances,
            pd.Series(future_balances)
        ]).reindex(report_periods, method="ffill")

        # Step 5. Formating

        # Create a new DataFrame with closing balances only
        balances = pd.DataFrame(
            [all_closing_balances],  # one row
            index=[0]  # index = 0
        )

        # Store balances for use in Graph
        self.graph_balances = all_closing_balances.values

        # Create a new DataFrame with repeated rows for each definition_id
        balances = pd.merge(definition_bal_df["definition_id"], balances, how='cross')
        balances = balances.set_index("definition_id")

        # ---------------------------------------------------------------------------------------------------------------------------------------
        # ********************** (4) Put together summarized accounts, totals and balances on end of each period *******************************
        # ---------------------------------------------------------------------------------------------------------------------------------------

        # Concatenate it all together
        report = pd.concat([pivot_cf, summarized_totals, balances])

        # -------------------------------------
        # Prepare report for visual appearance
        # -------------------------------------

        # Merge report definition header with report
        report = pd.merge(definition_df, report, left_on="definition_id", right_on="definition_id", how="left")

        # Sort based on key value
        report.sort_values("key", inplace=True)

        # Subtract report formatting in separate dataframe
        format_df = report["definition_type"]

        # Prepare report_df for visual appearance

        report.drop(columns=["definition_id", "key", "definition_type"], inplace=True)
        report.set_index("name", inplace=True)

        # Format column headers to show only the date part
        report.columns = [col.strftime(date_format()) if not pd.isnull(col) else col for col in report.columns]

        self.FORMAT = format_df
        return report



    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal:
            if role == Qt.ItemDataRole.DisplayRole:
                value = self.DATA.columns[section]
                return self._format_value(value, role)
            return None
        elif orientation == Qt.Orientation.Vertical:
            if role == Qt.ItemDataRole.DisplayRole:
                value = self.DATA.index[section]
                return self._format_value(value, role)
            # This works on Linux, but does not work on Windows
            # elif role == Qt.ItemDataRole.BackgroundRole:
            #     if self.FORMAT.iloc[section] == 2:  # Totals
            #         return QBrush(Qt.GlobalColor.darkBlue)
            #     elif self.FORMAT.iloc[section] == 3: # Balances
            #         return QBrush(Qt.GlobalColor.darkCyan)
            #     return None
            return None
        return None

    def _format_background(self, index, role=Qt.ItemDataRole.BackgroundRole):
        def_type = self.FORMAT.iloc[index.row()]
        if role == Qt.ItemDataRole.BackgroundRole:
            if def_type == 2: # Totals
                # Good for dark theme
                # return QBrush(Qt.GlobalColor.darkBlue)
                return QBrush(QColor("#25a5f5"))
            elif def_type == 3: # Balances
                return QBrush(Qt.GlobalColor.darkCyan)
            return None

        return None

class CheckingReportTable(ATable):
    def __init__(self, requery_method, parent=None):
        super().__init__(parent)
        self.requery_method = requery_method

    def action_requery(self):
        self.requery_method()

class CheckingReportModel(ATableModel):
    def __init__(self, table, engine, cashflow_model):
        super().__init__(table, engine)
        self.cashflow_model = cashflow_model


    def _do_requery(self):
        df = self.cashflow_model.checking_report
        df.columns = ["Periods", "NP ID", "NP nosaukums", "Fakts ieņēmumi", "Fakts maksājumi",
                                        "Sagaidāmie ieņēmumi", "Sagaidāmie maksājumi",
                                        "Budžeta ieņēmumi", "Budžeta maksājumi", "Fakts plus sagaidāmie ieņēmumi",
                                        "Fakts plus sagaidāmie maksājumi",
                                        "Ieņēmumi", "Maksājumi", "Neto naudas plūsma"]

        return df

